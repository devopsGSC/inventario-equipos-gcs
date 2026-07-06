import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Equipos"

navy  = "0F1724"
blue  = "3B82F6"
white = "FFFFFF"
amber = "FEF3C7"
thin  = Side(style='thin', color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Fila 1: Título
ws.merge_cells("A1:L1")
ws["A1"] = "InventarioTI — Plantilla de Actualización Masiva de Equipos"
ws["A1"].font = Font(name="Arial", bold=True, size=13, color=white)
ws["A1"].fill = PatternFill("solid", fgColor=navy)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Fila 2: Instrucción
ws.merge_cells("A2:L2")
ws["A2"] = "NumeroSerie es obligatorio (identifica el equipo). Los campos vacíos NO se modifican — solo se actualizan los que tengan valor."
ws["A2"].font = Font(name="Arial", size=9, color="92400E")
ws["A2"].fill = PatternFill("solid", fgColor="FEF3C7")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 20

# Fila 3: Separador
ws.row_dimensions[3].height = 6

# Fila 4: Encabezados
headers = [
    ("NumeroSerie *",  "A", 22),
    ("NombreEquipo",   "B", 28),
    ("TipoEquipo",     "C", 18),
    ("Marca",          "D", 18),
    ("Modelo",         "E", 22),
    ("IMEI",           "F", 20),
    ("Accesorios",     "G", 25),
    ("Costo",          "H", 12),
    ("FechaCompra",    "I", 16),
    ("FechaGarantia",  "J", 16),
    ("Estado",         "K", 16),
    ("Observaciones",  "L", 30),
]

for i, (title, col, width) in enumerate(headers, 1):
    cell = ws.cell(row=4, column=i, value=title)
    cell.font      = Font(name="Arial", bold=True, size=10, color=white)
    cell.fill      = PatternFill("solid", fgColor=blue)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = border
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[4].height = 22

# Filas 5-24: Datos con formato zebra
for row in range(5, 25):
    bg = "FFFFFF" if row % 2 == 0 else "F8FAFC"
    for col in range(1, 13):
        cell = ws.cell(row=row, column=col)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = border
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="center")
    # Formato fecha columnas I y J
    ws.cell(row=row, column=9).number_format  = "DD/MM/YYYY"
    ws.cell(row=row, column=10).number_format = "DD/MM/YYYY"
    # Formato número columna H
    ws.cell(row=row, column=8).number_format  = "#,##0.00"
    ws.row_dimensions[row].height = 18

# Validación desplegable para Estado (columna K) — usando fórmula de lista directa
dv = DataValidation(
    type="list",
    formula1='"Bodega,Asignado,Prestamo,EnGarantia,Baja"',
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Estado inválido",
    error="Use: Bodega, Asignado, Prestamo, EnGarantia o Baja"
)
ws.add_data_validation(dv)
dv.sqref = "K5:K24"

# Validación para TipoEquipo (columna C) desde hoja Referencia
dv_tipo = DataValidation(
    type="list",
    formula1="Referencia!$A$2:$A$6",
    allow_blank=True,
    showErrorMessage=False  # No bloquear — puede haber tipos personalizados
)
ws.add_data_validation(dv_tipo)
dv_tipo.sqref = "C5:C24"

# Hoja Referencia
ref = wb.create_sheet("Referencia")
ref["A1"] = "Tipos de equipo"
ref["A1"].font = Font(name="Arial", bold=True, size=10, color=white)
ref["A1"].fill = PatternFill("solid", fgColor=navy)
ref.column_dimensions["A"].width = 20
ref.column_dimensions["B"].width = 25

tipos = ["Laptop", "Teléfono", "Tablet", "Otro"]
for i, t in enumerate(tipos, 2):
    ref.cell(row=i, column=1, value=t).font = Font(name="Arial", size=10)
    ref.cell(row=i, column=1).border = border

ref["A7"]  = "Estados válidos"
ref["A7"].font = Font(name="Arial", bold=True, size=10, color=white)
ref["A7"].fill = PatternFill("solid", fgColor=navy)
estados = ["Bodega", "Asignado", "Prestamo", "EnGarantia", "Baja"]
for i, e in enumerate(estados, 8):
    ref.cell(row=i, column=1, value=e).font = Font(name="Arial", size=10)
    ref.cell(row=i, column=1).border = border

ref["A14"] = "Formato fechas:"
ref["A14"].font = Font(name="Arial", bold=True, size=10)
ref["A15"] = "DD/MM/YYYY"
ref["A15"].font = Font(name="Arial", size=10)
ref["A16"] = "Ej: 15/06/2025"
ref["A16"].font = Font(name="Arial", size=10, color="6B7280")

ref["A18"] = "Nota sobre campos vacíos:"
ref["A18"].font = Font(name="Arial", bold=True, size=10)
ref["B18"] = "Si una celda está vacía, ese campo NO se modifica en el sistema."
ref["B18"].font = Font(name="Arial", size=10, color="6B7280")
ref["B18"].alignment = Alignment(wrap_text=True)
ref.row_dimensions[18].height = 30

# Congelar fila de encabezados
ws.freeze_panes = "A5"

# Guardar en wwwroot del proyecto
import os
ruta = os.path.join(os.path.dirname(__file__), "wwwroot", "plantilla_actualizacion_equipos.xlsx")
wb.save(ruta)
print(f"Plantilla generada en: {ruta}")
